import json
from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side, Font
from style.column_style import decoration_nicname, design_of_amount_column, header_design
from typing import Dict


def table_layout_goal_and_assist(table_goal_and_assist: Worksheet) -> None:
    """
    Метод оформляет таблицу Goal and Assist.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_goal_and_assist: Worksheet - таблица Goal and Assist
    :return: None
    """
    header_design(name_table=table_goal_and_assist)
    decoration_nicname(name_table=table_goal_and_assist)

    with open('responses.json', encoding='utf-8') as file_coefficients:
        coefficients_goal_and_assist: Dict[str, float] = json.load(file_coefficients).get(
            'coefficients_goal_and_assist'
        )

    design_of_amount_column(table=table_goal_and_assist, coefficients=coefficients_goal_and_assist)

    style_border: Side = Side(border_style='thin', color='444444')

    for row in range(2, table_goal_and_assist.max_row + 1):
        for col in range(2, table_goal_and_assist.max_column + 1):
            cell: Cell = table_goal_and_assist.cell(column=col, row=row)
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            if cell.value == '+':
                cell.fill = PatternFill('solid', fgColor='55AA00')

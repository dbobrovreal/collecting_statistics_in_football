from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side
from typing import Dict


def design_of_amount_column(table: Worksheet, coefficients: Dict[str, float]) -> None:
    """
    Метод закрашивает колонку с итоговыми результатоми
    :param table: Worksheet - таблица в которой происходит заливка цветом
    :param coefficients: Dict[str, float] - словарь с коэффициентами
    :return: None
    """
    style: Side = Side(border_style='thin', color='444444')
    for num_row in range(2, table.max_row + 1):
        cell_sum: Cell = table.cell(row=num_row, column=table.max_column)
        cell_sum.border = Border(left=style, right=style, top=style, bottom=style)
        if cell_sum.value == 0:
            cell_sum.fill = PatternFill('solid', fgColor='FFAB91')
        elif cell_sum.value == coefficients.get('for_one'):
            cell_sum.fill = PatternFill('solid', fgColor='99BBFF')
        elif cell_sum.value == coefficients.get('in_two'):
            cell_sum.fill = PatternFill('solid', fgColor='5599FF')
        elif cell_sum.value == coefficients.get('for_three'):
            cell_sum.fill = PatternFill('solid', fgColor='0066FF')
        elif cell_sum.value == coefficients.get('for_four'):
            cell_sum.fill = PatternFill('solid', fgColor='0044BB')


def decoration_nicname(name_table: Worksheet) -> None:
    """
    Метод оформляет столбец с Nicname игроков
    :param name_table: Worksheet - Название листа в таблице excel
    :return: None
    """
    name_table.column_dimensions['A'].width = 15
    style: Side = Side(border_style='medium', color='444444')
    for num_row in range(1, name_table.max_row + 1):
        cell: Cell = name_table.cell(column=1, row=num_row)
        cell.border = Border(left=style, right=style, top=style, bottom=style)


def header_design(name_table: Worksheet) -> None:
    """
    Метод оформляет шапку таблицы
    :param name_table: Worksheet - Название листа в таблице excel
    :return: None
    """
    style_hats: Side = Side(border_style='medium', color='444444')
    for num_column in range(1, name_table.max_column + 1):
        cell: Cell = name_table.cell(row=1, column=num_column)
        cell.border = Border(left=style_hats, right=style_hats, top=style_hats, bottom=style_hats)


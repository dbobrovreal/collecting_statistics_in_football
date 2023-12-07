from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side, Font


def design_of_manager_table(table_manager: Worksheet) -> None:
    """
    Метод оформляет таблицу Manager.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_manager: Worksheet - Таблица Manager
    :return: None
    """
    table_manager.column_dimensions['A'].width = 20
    table_manager.column_dimensions['D'].width = 20

    column_result: tuple = table_manager['D']

    filling_win: PatternFill = PatternFill('solid', fgColor='0000FF00')
    filling_draw: PatternFill = PatternFill('solid', fgColor='FFD700')
    filling_def: PatternFill = PatternFill('solid', fgColor='00FF0000')
    for column in column_result:
        number_column: int = column.row
        cell: Cell = table_manager.cell(column=4, row=number_column)
        style_border: Side = Side(border_style='thin', color='444444')
        cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
        cell.font = Font(bold=True)
        match cell.value:
            case 'win':
                cell.fill = filling_win
            case 'draw':
                cell.fill = filling_draw
            case 'def':
                cell.fill = filling_def

from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side
from typing import List


def table_layout_example(table_example: Worksheet) -> None:
    """
    Метод оформляет таблицу Worksheet
    :param table_example: Worksheet - Таблица Example
    :return: None
    """
    list_of_columns: List[str] = ['A', 'B', 'C', 'D', 'E', 'F']

    for column in list_of_columns:
        table_example.column_dimensions[column].width = 15

    fill = PatternFill('solid', fgColor='9FA8DA')
    style_border: Side = Side(border_style='thin', color='444444')
    for row in range(2, table_example.max_row + 1):
        if row == 2:
            fill = PatternFill('solid', fgColor='283593')
        elif 2 < row <= 5:
            fill = PatternFill('solid', fgColor='3F51B5')
        elif 5 < row <= 8:
            fill = PatternFill('solid', fgColor='7986CB')
        elif row > 8:
            fill = PatternFill('solid', fgColor='C5CAE9')

        cell: Cell = table_example.cell(row=row, column=6)
        cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
        cell.fill = fill

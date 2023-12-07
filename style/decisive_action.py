from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side, Font
from typing import List


def table_layout_decisive_action(table_decisive_action: Worksheet) -> None:
    """
    Метод оформляет таблицу Decisive Action.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_decisive_action: Worksheet - Таблица Decisive Action
    :return: None
    """
    list_of_columns: List[str] = ['A', 'C', 'E', 'F']
    for column in list_of_columns:
        table_decisive_action.column_dimensions[column].width = 15

    column_sum_pst: tuple = table_decisive_action['J']

    fill_color: PatternFill = PatternFill('solid', fgColor='FFD700')
    style_hats: Side = Side(border_style='medium', color='444444')

    for num_clumn in range(1, table_decisive_action.max_column + 1):
        point_cell = table_decisive_action.cell(row=2, column=num_clumn)
        point_cell.fill = fill_color
        point_cell.border = Border(left=style_hats, right=style_hats, top=style_hats, bottom=style_hats)

    for cell_column in column_sum_pst[2:]:
        cell: Cell = table_decisive_action.cell(column=10, row=cell_column.row)
        style_border: Side = Side(border_style='thin', color='444444')
        cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
        cell.font = Font(bold=True)
        if cell.value < 15:
            cell.fill = PatternFill('solid', fgColor='99BBFF')
        elif 15 < cell.value < 25:
            cell.fill = PatternFill('solid', fgColor='5599FF')
        elif 25 < cell.value < 35:
            cell.fill = PatternFill('solid', fgColor='0066FF')
        elif 35 < cell.value < 45:
            cell.fill = PatternFill('solid', fgColor='0044BB')
        else:
            cell.fill = PatternFill('solid', fgColor='0000FF')

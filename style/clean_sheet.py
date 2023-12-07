import json
from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side, Font
from style.column_style import decoration_nicname, design_of_amount_column, header_design
from typing import Dict


def table_layout_clear_sheet(table_clear_sheet: Worksheet) -> None:
    """
    Метод оформляет таблицу Clear Sheet.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_clear_sheet: Worksheet - Таблица Clear Sheet
    :return: None
    """
    header_design(name_table=table_clear_sheet)
    decoration_nicname(name_table=table_clear_sheet)

    with open('responses.json', encoding='utf-8') as file_coefficients:
        coefficients_clean_sheet: Dict[str, float] = json.load(file_coefficients).get('coefficients_clean_sheet')

    design_of_amount_column(table=table_clear_sheet, coefficients=coefficients_clean_sheet)

    for row in range(2, table_clear_sheet.max_row + 1):
        for col in range(2, table_clear_sheet.max_column):
            cell: Cell = table_clear_sheet.cell(column=col, row=row)
            style_border: Side = Side(border_style='thin', color='444444')
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            match cell.value:
                case 0:
                    cell.fill = PatternFill('solid', fgColor='00FF0000')
                case 1:
                    cell.fill = PatternFill('solid', fgColor='55AA00')

from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Border, Side, Font
from style.column_style import decoration_nicname, header_design


def table_layout_goal(table_goal: Worksheet) -> None:
    """
    Метод оформляет таблицу Goal.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_goal: Worksheet - Таблица Goal
    :return: None
    """
    header_design(name_table=table_goal)
    decoration_nicname(name_table=table_goal)

    for row in range(2, table_goal.max_row + 1):
        for col in range(2, table_goal.max_column):
            cell: Cell = table_goal.cell(column=col, row=row)
            style_border: Side = Side(border_style='thin', color='444444')
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            if cell.value == 0:
                cell.fill = PatternFill('solid', fgColor='00FF0000')
            else:
                cell.fill = PatternFill('solid', fgColor='55AA00')

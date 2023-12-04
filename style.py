import openpyxl
from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from typing import List


def header_design(name_table: Worksheet) -> None:
    """
    Метод оформляет шапку таблицы
    :param name_table: Worksheet - Название листа в таблице excel
    :return: None
    """
    for num_column in range(1, name_table.max_column + 1):
        cell: Cell = name_table.cell(row=1, column=num_column)
        style_hats: Side = Side(border_style='medium', color='444444')
        cell.border = Border(left=style_hats, right=style_hats, top=style_hats, bottom=style_hats)


def decoration_nicname(name_table: Worksheet) -> None:
    """
    Метод оформляет столбец с Nicname игроков
    :param name_table: Worksheet - Название листа в таблице excel
    :return: None
    """
    name_table.column_dimensions['A'].width = 15

    for num_row in range(1, name_table.max_row + 1):
        cell: Cell = name_table.cell(column=1, row=num_row)
        style: Side = Side(border_style='medium', color='444444')
        cell.border = Border(left=style, right=style, top=style, bottom=style)


def set_column_sizes_players_statistical_table(table_statistic: Worksheet) -> None:
    """
    Метод оформляет таблицу Player Statistic.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_statistic: Worksheet - таблица Player Statistic
    :return: None
    """
    table_statistic.column_dimensions['A'].width = 20
    list_of_columns: List[str] = ['E', 'F', 'G', 'I', 'J', 'K', 'L']

    for column in list_of_columns:
        table_statistic.column_dimensions[column].width = 15


def design_of_manager_table(table_manager: Worksheet) -> None:
    """
    Метод оформляет таблицу Manager.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_manager: Worksheet - Таблица Manager
    :return: None
    """
    table_manager.column_dimensions['A'].width = 20
    table_manager.column_dimensions['D'].width = 20

    column_result: tuple = table_manager['D']

    filling_win: PatternFill = PatternFill('solid', fgColor='0000FF00')
    filling_draw: PatternFill = PatternFill('solid', fgColor='FFD700')
    filling_def: PatternFill = PatternFill('solid', fgColor='00FF0000')
    for column in column_result:
        number_column: int = column.row
        cell: Cell = table_manager.cell(column=4, row=number_column)
        style_border: Side = Side(border_style='thin', color='444444')
        cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
        cell.font = Font(bold=True)
        match cell.value:
            case 'win':
                cell.fill = filling_win
            case 'draw':
                cell.fill = filling_draw
            case 'def':
                cell.fill = filling_def


def table_layout_decisive_action(table_decisive_action: Worksheet) -> None:
    """
    Метод оформляет таблицу Decisive Action.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_decisive_action: Worksheet - Таблица Decisive Action
    :return: None
    """
    list_of_columns: List[str] = ['A', 'C', 'E', 'F']
    for column in list_of_columns:
        table_decisive_action.column_dimensions[column].width = 15

    column_sum_pst: tuple = table_decisive_action['J']

    for cell_column in column_sum_pst[1:]:
        cell: Cell = table_decisive_action.cell(column=10, row=cell_column.row)
        style_border: Side = Side(border_style='thin', color='444444')
        cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
        cell.font = Font(bold=True)
        if cell.value < 15:
            cell.fill = PatternFill('solid', fgColor='99BBFF')
        elif 15 < cell.value < 25:
            cell.fill = PatternFill('solid', fgColor='5599FF')
        elif 25 < cell.value < 35:
            cell.fill = PatternFill('solid', fgColor='0066FF')
        elif 35 < cell.value < 45:
            cell.fill = PatternFill('solid', fgColor='0044BB')
        else:
            cell.fill = PatternFill('solid', fgColor='0000FF')


def table_layout_goal(table_goal: Worksheet) -> None:
    """
    Метод оформляет таблицу Goal.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_goal: Worksheet - Таблица Goal
    :return: None
    """
    header_design(name_table=table_goal)
    decoration_nicname(name_table=table_goal)

    for row in range(2, table_goal.max_row + 1):
        for col in range(2, table_goal.max_column + 1):
            cell: Cell = table_goal.cell(column=col, row=row)
            style_border: Side = Side(border_style='thin', color='444444')
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            match cell.value:
                case 0:
                    cell.fill = PatternFill('solid', fgColor='00FF0000')
                case 3:
                    cell.fill = PatternFill('solid', fgColor='55AA00')


def table_layout_clear_sheet(table_clear_sheet: Worksheet) -> None:
    """
    Метод оформляет таблицу Clear Sheet.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_clear_sheet: Worksheet - Таблица Clear Sheet
    :return: None
    """
    header_design(name_table=table_clear_sheet)
    decoration_nicname(name_table=table_clear_sheet)

    for row in range(2, table_clear_sheet.max_row + 1):
        for col in range(2, table_clear_sheet.max_column + 1):
            cell: Cell = table_clear_sheet.cell(column=col, row=row)
            style_border: Side = Side(border_style='thin', color='444444')
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            match cell.value:
                case 0:
                    cell.fill = PatternFill('solid', fgColor='00FF0000')
                case 1:
                    cell.fill = PatternFill('solid', fgColor='55AA00')


def table_layout_goal_and_assist(table_goal_and_assist: Worksheet) -> None:
    """
    Метод оформляет таблицу Goal and Assist.
    Задает размер колонки, окрашивает нужные ячейки
    и выделяет границы полей ячейки
    :param table_goal_and_assist: Worksheet - таблица Goal and Assist
    :return: None
    """
    header_design(name_table=table_goal_and_assist)
    decoration_nicname(name_table=table_goal_and_assist)

    for row in range(2, table_goal_and_assist.max_row + 1):
        for col in range(2, table_goal_and_assist.max_column + 1):
            cell: Cell = table_goal_and_assist.cell(column=col, row=row)
            style_border: Side = Side(border_style='thin', color='444444')
            cell.border = Border(left=style_border, right=style_border, top=style_border, bottom=style_border)
            cell.font = Font(bold=True)
            if cell.value == '+':
                cell.fill = PatternFill('solid', fgColor='55AA00')


def setting_style_for_the_tables(path_file: str) -> None:
    """
    Метод оформляет стили для каждой таблице excel с итоговой информацией
    :param path_file: str - путь до файла с итоговыми результатоми
    :return: None
    """
    file_tables: Workbook = openpyxl.load_workbook(path_file)

    player_statistics: Worksheet = file_tables['Player statistic']
    set_column_sizes_players_statistical_table(table_statistic=player_statistics)

    manager_statistic: Worksheet = file_tables['Manager']
    design_of_manager_table(table_manager=manager_statistic)

    decisive_action: Worksheet = file_tables['Decisive action']
    table_layout_decisive_action(table_decisive_action=decisive_action)

    goal: Worksheet = file_tables['Goal']
    table_layout_goal(table_goal=goal)

    clear_sheet: Worksheet = file_tables['Clear sheet']
    table_layout_clear_sheet(table_clear_sheet=clear_sheet)

    goal_and_assist: Worksheet = file_tables['Goal and Assist']
    table_layout_goal_and_assist(table_goal_and_assist=goal_and_assist)

    file_tables.save(path_file)
